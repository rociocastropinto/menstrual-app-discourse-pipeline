"""
sample_100_for_iaa.py
---------------------
Draws a stratified random sample of 100 sentences from 06_annotated_2673.csv
for inter-annotator agreement (Cohen's Kappa).

Stratification is by year so all temporal periods are represented.
Output: iaa_sample.xlsx  — two sheets:
  - Sheet 1 "sentences_to_annotate": sentence + LLM labels hidden (for your manual annotation)
  - Sheet 2 "llm_labels_hidden":     LLM labels stored separately (reveal after you annotate)

Usage:
    python sample_100_for_iaa.py
    # Edit: set INPUT_CSV to the path of your 06_annotated_2673.csv
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
INPUT_CSV   = "06_annotated_2673.csv"   # path to your annotated corpus
OUTPUT_XLSX = "iaa_sample.xlsx"
N_SAMPLE    = 100
RANDOM_SEED = 42
# ─────────────────────────────────────────────────────────────────────────────

df = pd.read_csv(INPUT_CSV)

# Normalise column names (strip whitespace)
df.columns = df.columns.str.strip()

# Extract year from date column (try multiple formats)
if "date" in df.columns:
    df["year"] = pd.to_datetime(df["date"], errors="coerce").dt.year
elif "year" in df.columns:
    pass
else:
    raise ValueError("No 'date' or 'year' column found. Check your CSV.")

# Drop rows with missing year or missing sentence
df = df.dropna(subset=["year", "sentence"])
df["year"] = df["year"].astype(int)

# ── STRATIFIED SAMPLING ───────────────────────────────────────────────────────
year_counts = df["year"].value_counts().sort_index()
print("Sentences per year in full corpus:")
print(year_counts.to_string())

# Proportional allocation, minimum 1 per year that has any sentences
proportions = year_counts / year_counts.sum()
allocations = (proportions * N_SAMPLE).round().astype(int)

# Adjust rounding so total == N_SAMPLE
diff = N_SAMPLE - allocations.sum()
# Add/subtract from the largest year
largest_year = allocations.idxmax()
allocations[largest_year] += diff

print(f"\nSample allocation per year (total = {allocations.sum()}):")
print(allocations.to_string())

sampled_frames = []
for year, n in allocations.items():
    year_df = df[df["year"] == year]
    n = min(n, len(year_df))  # can't sample more than available
    sampled_frames.append(year_df.sample(n=n, random_state=RANDOM_SEED))

sample = pd.concat(sampled_frames).reset_index(drop=True)
sample.insert(0, "iaa_id", range(1, len(sample) + 1))

print(f"\nFinal sample: {len(sample)} sentences")

# ── ANNOTATION COLUMNS (LLM labels) ──────────────────────────────────────────
llm_cols = [
    "construction_type",
    "actor_category",
    "responsibility_direction",
    "agent_suppressed",
]

# Verify all LLM label columns exist
for col in llm_cols:
    if col not in sample.columns:
        print(f"WARNING: column '{col}' not found in CSV — will be blank in output")

# ── BUILD EXCEL OUTPUT ────────────────────────────────────────────────────────
wb = Workbook()

# ── SHEET 1: For manual annotation (NO LLM labels visible) ───────────────────
ws1 = wb.active
ws1.title = "annotate_here"

headers_annotate = [
    "iaa_id", "year", "article_id", "sentence",
    "YOUR_construction_type",
    "YOUR_actor_category",
    "YOUR_responsibility_direction",
    "YOUR_agent_suppressed",
    "notes"
]

ws1.append(headers_annotate)

# Style header
h_fill = PatternFill("solid", start_color="2E4057")
h_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
for cell in ws1[1]:
    cell.fill = h_fill
    cell.font = h_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

input_fill  = PatternFill("solid", start_color="FFF9C4")   # yellow = cells to fill in
normal_fill_a = PatternFill("solid", start_color="F5F5F5")
normal_fill_b = PatternFill("solid", start_color="FFFFFF")

for i, row in sample.iterrows():
    excel_row = i + 2
    fill = normal_fill_a if excel_row % 2 == 0 else normal_fill_b
    data = [
        row["iaa_id"],
        row.get("year", ""),
        row.get("article_id", ""),
        row.get("sentence", ""),
        "",  # YOUR_construction_type
        "",  # YOUR_actor_category
        "",  # YOUR_responsibility_direction
        "",  # YOUR_agent_suppressed
        "",  # notes
    ]
    ws1.append(data)
    for j, cell in enumerate(ws1[excel_row], start=1):
        if j >= 5:  # annotation columns
            cell.fill = input_fill
            cell.font = Font(name="Arial", size=10)
        else:
            cell.fill = fill
            cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# Column widths sheet 1
col_widths_1 = [6, 6, 14, 80, 22, 22, 26, 18, 20]
for i, w in enumerate(col_widths_1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.row_dimensions[1].height = 28
for i in range(2, len(sample) + 2):
    ws1.row_dimensions[i].height = 60

ws1.freeze_panes = "A2"

# Add dropdown validation note in a comment row at top
ws1.insert_rows(1)
ws1["A1"] = (
    "INSTRUCTIONS — Fill in columns E–H only. "
    "construction_type: active_transitive | passive | nominalization | intransitive | other  //  "
    "actor_category: corporate | regulatory | user | technological_system | expert_civil_society | suppressed | other  //  "
    "responsibility_direction: structural | individualising | diffuse | neutral  //  "
    "agent_suppressed: true | false"
)
ws1["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
ws1["A1"].fill = PatternFill("solid", start_color="C0392B")
ws1["A1"].alignment = Alignment(wrap_text=True, vertical="center")
ws1.merge_cells("A1:I1")
ws1.row_dimensions[1].height = 40

# ── SHEET 2: LLM labels (sealed reference — reveal after you annotate) ────────
ws2 = wb.create_sheet("llm_labels_REVEAL_AFTER")

headers_llm = [
    "iaa_id", "year", "article_id", "sentence",
    "LLM_construction_type",
    "LLM_actor_category",
    "LLM_responsibility_direction",
    "LLM_agent_suppressed",
    "LLM_confidence",
]

ws2.append(headers_llm)
for cell in ws2[1]:
    cell.fill = h_fill
    cell.font = h_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for i, row in sample.iterrows():
    excel_row = i + 2
    fill = normal_fill_a if excel_row % 2 == 0 else normal_fill_b
    data = [
        row["iaa_id"],
        row.get("year", ""),
        row.get("article_id", ""),
        row.get("sentence", ""),
        row.get("construction_type", ""),
        row.get("actor_category", ""),
        row.get("responsibility_direction", ""),
        str(row.get("agent_suppressed", "")).lower(),
        row.get("confidence", ""),
    ]
    ws2.append(data)
    for cell in ws2[excel_row]:
        cell.fill = fill
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

col_widths_2 = [6, 6, 14, 80, 22, 22, 26, 18, 12]
for i, w in enumerate(col_widths_2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.row_dimensions[1].height = 28
for i in range(2, len(sample) + 2):
    ws2.row_dimensions[i].height = 60

ws2.freeze_panes = "A2"

# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_XLSX)
print(f"\nSaved: {OUTPUT_XLSX}")
print("Sheet 1 'annotate_here'         — fill in columns E–H manually")
print("Sheet 2 'llm_labels_REVEAL_AFTER' — open only after you finish annotating")